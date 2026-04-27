# 处理 NGSIM 数据，筛选出 Location 列为 "US-101" 的行，并删除指定的列
from pathlib import Path

from openpyxl import load_workbook, Workbook

DATASET_DIR = Path(__file__).resolve().parents[2]
input_file = DATASET_DIR / "raw" / "original_dataset" / "NGSIM.xlsx"
output_file = DATASET_DIR / "intermediate" / "us101_filtered.xlsx"

columns_to_drop = {"O_Zone", "D_Zone", "Int_ID", "Section_ID", "Direction", "Movement"}

wb = load_workbook(input_file, read_only=True, data_only=True)
ws = wb.active

out_wb = Workbook(write_only=True)
out_ws = out_wb.create_sheet()

rows = ws.iter_rows(values_only=True)
header = next(rows)
header = [str(x) if x is not None else "" for x in header]

if "Location" not in header:
    raise ValueError("找不到 Location 列")

location_idx = header.index("Location")
keep_indices = [i for i, col in enumerate(header) if col not in columns_to_drop]
output_header = [header[i] for i in keep_indices]

out_ws.append(output_header)

for row in rows:
    location_value = row[location_idx]
    if str(location_value).strip().lower() == "us-101":
        out_ws.append([row[i] for i in keep_indices])

output_file.parent.mkdir(parents=True, exist_ok=True)
out_wb.save(output_file)

print(f"处理完成，结果已保存到 {output_file}")

# import zipfile

# path = r"NGSIM.csv"

# with zipfile.ZipFile(path, "r") as z:
#     print(z.namelist())
