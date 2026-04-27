from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook


REQUIRED_COLUMNS = [
    "Vehicle_ID",
    "Global_Time",
    "v_Vel",
    "v_Acc",
    "Time_Headway",
    "Space_Headway",
]


@dataclass
class Config:
    input_path: str
    output_path: str
    sheet_name: Optional[str]
    window_seconds: int
    histogram_min: float
    histogram_max: float
    histogram_bin_width: float
    free_flow_speed: Optional[float]
    free_flow_quantile: float
    low_speed_ratio_to_free_flow: float
    near_stop_threshold: float
    hard_brake_threshold: float
    speed_drop_threshold: float
    short_time_headway_threshold: float
    min_points_per_vehicle: int


class VehicleWindowAgg:
    __slots__ = (
        "count",
        "speed_sum",
        "speed_sq_sum",
        "speed_min",
        "speed_first",
        "speed_last",
        "acc_sum",
        "acc_sq_sum",
        "acc_min",
        "hard_brake_flag",
        "time_headway_sum",
        "time_headway_count",
        "space_headway_sum",
        "space_headway_count",
    )

    def __init__(self, speed: float, acc: float, time_headway: Optional[float], space_headway: Optional[float], hard_brake_threshold: float):
        self.count = 1
        self.speed_sum = speed
        self.speed_sq_sum = speed * speed
        self.speed_min = speed
        self.speed_first = speed
        self.speed_last = speed
        self.acc_sum = acc
        self.acc_sq_sum = acc * acc
        self.acc_min = acc
        self.hard_brake_flag = acc < hard_brake_threshold
        if time_headway is not None and time_headway > 0:
            self.time_headway_sum = time_headway
            self.time_headway_count = 1
        else:
            self.time_headway_sum = 0.0
            self.time_headway_count = 0
        if space_headway is not None and space_headway > 0:
            self.space_headway_sum = space_headway
            self.space_headway_count = 1
        else:
            self.space_headway_sum = 0.0
            self.space_headway_count = 0

    def add(self, speed: float, acc: float, time_headway: Optional[float], space_headway: Optional[float], hard_brake_threshold: float) -> None:
        self.count += 1
        self.speed_sum += speed
        self.speed_sq_sum += speed * speed
        if speed < self.speed_min:
            self.speed_min = speed
        self.speed_last = speed
        self.acc_sum += acc
        self.acc_sq_sum += acc * acc
        if acc < self.acc_min:
            self.acc_min = acc
        if acc < hard_brake_threshold:
            self.hard_brake_flag = True
        if time_headway is not None and time_headway > 0:
            self.time_headway_sum += time_headway
            self.time_headway_count += 1
        if space_headway is not None and space_headway > 0:
            self.space_headway_sum += space_headway
            self.space_headway_count += 1

    def finalize(self, low_speed_threshold: float, near_stop_threshold: float, speed_drop_threshold: float, short_time_headway_threshold: float, min_points_per_vehicle: int) -> Optional[Tuple[float, float, float, bool, bool, float, float, float, float, bool, Optional[float], Optional[float], bool]]:
        if self.count < min_points_per_vehicle:
            return None
        speed_mean = self.speed_sum / self.count
        speed_min = self.speed_min
        speed_delta = self.speed_last - self.speed_first
        low_speed_flag = speed_mean < low_speed_threshold
        near_stop_flag = speed_min < near_stop_threshold
        acc_mean = self.acc_sum / self.count
        acc_min = self.acc_min
        hard_brake_flag = self.hard_brake_flag
        speed_drop_flag = speed_delta < -speed_drop_threshold
        if self.time_headway_count > 0:
            time_headway_mean = self.time_headway_sum / self.time_headway_count
            short_headway_flag = time_headway_mean < short_time_headway_threshold
        else:
            time_headway_mean = None
            short_headway_flag = False
        if self.space_headway_count > 0:
            space_headway_mean = self.space_headway_sum / self.space_headway_count
        else:
            space_headway_mean = None
        return (
            speed_mean,
            speed_min,
            speed_delta,
            low_speed_flag,
            near_stop_flag,
            acc_mean,
            acc_min,
            self.speed_first,
            self.speed_last,
            hard_brake_flag,
            time_headway_mean,
            space_headway_mean,
            short_headway_flag,
        )


def safe_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def find_sheet(workbook, sheet_name: Optional[str]):
    if sheet_name:
        return workbook[sheet_name]
    return workbook.active


def header_map_from_sheet(sheet) -> Dict[str, int]:
    rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows)
    mapping: Dict[str, int] = {}
    for idx, name in enumerate(header):
        if name is not None:
            mapping[str(name).strip()] = idx
    missing = [c for c in REQUIRED_COLUMNS if c not in mapping]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    return mapping


def percentile(values: Sequence[float], q: float) -> Optional[float]:
    if not values:
        return None
    if len(values) == 1:
        return float(values[0])
    sorted_vals = sorted(values)
    pos = (len(sorted_vals) - 1) * q
    lower = int(math.floor(pos))
    upper = int(math.ceil(pos))
    if lower == upper:
        return float(sorted_vals[lower])
    weight = pos - lower
    return float(sorted_vals[lower] * (1.0 - weight) + sorted_vals[upper] * weight)


def mean(values: Sequence[float]) -> Optional[float]:
    if not values:
        return None
    return float(sum(values) / len(values))


def std(values: Sequence[float]) -> Optional[float]:
    if not values:
        return None
    if len(values) == 1:
        return 0.0
    m = sum(values) / len(values)
    return float(math.sqrt(sum((v - m) * (v - m) for v in values) / len(values)))


def to_utc_iso(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000.0, tz=timezone.utc).isoformat()


def compute_window_row(
    window_id: int,
    base_time_ms: int,
    window_ms: int,
    vehicle_map: Dict[int, VehicleWindowAgg],
    low_speed_threshold: float,
    near_stop_threshold: float,
    speed_drop_threshold: float,
    short_time_headway_threshold: float,
    min_points_per_vehicle: int,
) -> Optional[List[object]]:
    speed_means: List[float] = []
    speed_mins: List[float] = []
    acc_means: List[float] = []
    acc_mins: List[float] = []
    low_speed_count = 0
    near_stop_count = 0
    hard_brake_count = 0
    speed_drop_count = 0
    stop_vehicle_count = 0
    time_headway_means: List[float] = []
    short_headway_count = 0
    space_headway_means: List[float] = []

    for agg in vehicle_map.values():
        item = agg.finalize(
            low_speed_threshold=low_speed_threshold,
            near_stop_threshold=near_stop_threshold,
            speed_drop_threshold=speed_drop_threshold,
            short_time_headway_threshold=short_time_headway_threshold,
            min_points_per_vehicle=min_points_per_vehicle,
        )
        if item is None:
            continue
        speed_mean_v, speed_min_v, speed_delta_v, low_speed_flag, near_stop_flag, acc_mean_v, acc_min_v, _, _, hard_brake_flag, time_headway_mean_v, space_headway_mean_v, short_headway_flag = item
        speed_means.append(speed_mean_v)
        speed_mins.append(speed_min_v)
        acc_means.append(acc_mean_v)
        acc_mins.append(acc_min_v)
        if low_speed_flag:
            low_speed_count += 1
        if near_stop_flag:
            near_stop_count += 1
            stop_vehicle_count += 1
        if hard_brake_flag:
            hard_brake_count += 1
        if speed_delta_v < -speed_drop_threshold:
            speed_drop_count += 1
        if time_headway_mean_v is not None:
            time_headway_means.append(time_headway_mean_v)
            if short_headway_flag:
                short_headway_count += 1
        if space_headway_mean_v is not None:
            space_headway_means.append(space_headway_mean_v)

    vehicle_count = len(speed_means)
    if vehicle_count == 0:
        return None

    valid_time_headway_count = len(time_headway_means)
    window_start_ms = base_time_ms + window_id * window_ms
    window_end_ms = window_start_ms + window_ms

    row: List[object] = [
        window_id,
        window_start_ms,
        window_end_ms,
        to_utc_iso(window_start_ms),
        to_utc_iso(window_end_ms),
        vehicle_count,
        mean(speed_means),
        std(speed_means),
        percentile(speed_means, 0.10),
        percentile(speed_means, 0.50),
        percentile(speed_means, 0.90),
        min(speed_mins) if speed_mins else None,
        low_speed_count / vehicle_count,
        near_stop_count / vehicle_count,
        stop_vehicle_count,
        mean(acc_means),
        std(acc_means),
        min(acc_mins) if acc_mins else None,
        hard_brake_count / vehicle_count,
        speed_drop_count / vehicle_count,
        mean(time_headway_means),
        percentile(time_headway_means, 0.10),
        (short_headway_count / valid_time_headway_count) if valid_time_headway_count > 0 else None,
        mean(space_headway_means),
        percentile(space_headway_means, 0.10),
    ]
    return row


def first_pass(config: Config) -> Tuple[int, float]:
    workbook = load_workbook(config.input_path, read_only=True, data_only=True)
    sheet = find_sheet(workbook, config.sheet_name)
    mapping = header_map_from_sheet(sheet)

    bin_width = config.histogram_bin_width
    bin_count = int(math.ceil((config.histogram_max - config.histogram_min) / bin_width))
    histogram = [0] * bin_count
    base_time_ms: Optional[int] = None
    valid_speed_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        global_time = safe_float(row[mapping["Global_Time"]])
        speed = safe_float(row[mapping["v_Vel"]])
        if global_time is None or speed is None:
            continue
        global_time_i = int(global_time)
        if base_time_ms is None or global_time_i < base_time_ms:
            base_time_ms = global_time_i
        clamped = min(max(speed, config.histogram_min), config.histogram_max - 1e-12)
        idx = int((clamped - config.histogram_min) / bin_width)
        histogram[idx] += 1
        valid_speed_count += 1

    workbook.close()

    if base_time_ms is None:
        raise ValueError("No valid rows found.")

    if config.free_flow_speed is not None:
        return base_time_ms, config.free_flow_speed

    if valid_speed_count == 0:
        raise ValueError("No valid speed values found.")

    target = valid_speed_count * config.free_flow_quantile
    running = 0
    estimated = config.histogram_max
    for idx, count in enumerate(histogram):
        running += count
        if running >= target:
            estimated = config.histogram_min + (idx + 0.5) * bin_width
            break

    return base_time_ms, estimated


def second_pass(config: Config, base_time_ms: int, free_flow_speed: float) -> None:
    workbook = load_workbook(config.input_path, read_only=True, data_only=True)
    sheet = find_sheet(workbook, config.sheet_name)
    mapping = header_map_from_sheet(sheet)

    output_wb = Workbook(write_only=True)
    ws = output_wb.create_sheet(title="window_5s_features")
    ws.append([
        "window_id",
        "window_start_ms",
        "window_end_ms",
        "window_start_utc",
        "window_end_utc",
        "win_vehicle_count",
        "win_speed_mean",
        "win_speed_std",
        "win_speed_p10",
        "win_speed_p50",
        "win_speed_p90",
        "win_speed_min",
        "win_low_speed_ratio",
        "win_near_stop_ratio",
        "win_stop_vehicle_count",
        "win_acc_mean",
        "win_acc_std",
        "win_acc_min",
        "win_hard_brake_ratio",
        "win_speed_drop_ratio",
        "win_time_headway_mean",
        "win_time_headway_p10",
        "win_short_headway_ratio",
        "win_space_headway_mean",
        "win_space_headway_p10",
    ])

    meta = output_wb.create_sheet(title="meta")
    meta.append(["key", "value"])
    meta.append(["input_path", config.input_path])
    meta.append(["window_seconds", config.window_seconds])
    meta.append(["base_time_ms", base_time_ms])
    meta.append(["free_flow_speed_ftps", free_flow_speed])
    meta.append(["low_speed_threshold_ftps", free_flow_speed * config.low_speed_ratio_to_free_flow])
    meta.append(["near_stop_threshold_ftps", config.near_stop_threshold])
    meta.append(["hard_brake_threshold_ftps2", config.hard_brake_threshold])
    meta.append(["speed_drop_threshold_ftps", config.speed_drop_threshold])
    meta.append(["short_time_headway_threshold_s", config.short_time_headway_threshold])
    meta.append(["min_points_per_vehicle", config.min_points_per_vehicle])

    window_ms = config.window_seconds * 1000
    low_speed_threshold = free_flow_speed * config.low_speed_ratio_to_free_flow

    active_window_id: Optional[int] = None
    vehicle_map: Dict[int, VehicleWindowAgg] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        vehicle_id = safe_float(row[mapping["Vehicle_ID"]])
        global_time = safe_float(row[mapping["Global_Time"]])
        speed = safe_float(row[mapping["v_Vel"]])
        acc = safe_float(row[mapping["v_Acc"]])
        time_headway = safe_float(row[mapping["Time_Headway"]])
        space_headway = safe_float(row[mapping["Space_Headway"]])

        if vehicle_id is None or global_time is None or speed is None or acc is None:
            continue

        window_id = int((int(global_time) - base_time_ms) // window_ms)

        if active_window_id is None:
            active_window_id = window_id

        if window_id != active_window_id:
            row_out = compute_window_row(
                window_id=active_window_id,
                base_time_ms=base_time_ms,
                window_ms=window_ms,
                vehicle_map=vehicle_map,
                low_speed_threshold=low_speed_threshold,
                near_stop_threshold=config.near_stop_threshold,
                speed_drop_threshold=config.speed_drop_threshold,
                short_time_headway_threshold=config.short_time_headway_threshold,
                min_points_per_vehicle=config.min_points_per_vehicle,
            )
            if row_out is not None:
                ws.append(row_out)
            vehicle_map = {}
            active_window_id = window_id

        vid = int(vehicle_id)
        if vid not in vehicle_map:
            vehicle_map[vid] = VehicleWindowAgg(
                speed=speed,
                acc=acc,
                time_headway=time_headway,
                space_headway=space_headway,
                hard_brake_threshold=config.hard_brake_threshold,
            )
        else:
            vehicle_map[vid].add(
                speed=speed,
                acc=acc,
                time_headway=time_headway,
                space_headway=space_headway,
                hard_brake_threshold=config.hard_brake_threshold,
            )

    if active_window_id is not None:
        row_out = compute_window_row(
            window_id=active_window_id,
            base_time_ms=base_time_ms,
            window_ms=window_ms,
            vehicle_map=vehicle_map,
            low_speed_threshold=low_speed_threshold,
            near_stop_threshold=config.near_stop_threshold,
            speed_drop_threshold=config.speed_drop_threshold,
            short_time_headway_threshold=config.short_time_headway_threshold,
            min_points_per_vehicle=config.min_points_per_vehicle,
        )
        if row_out is not None:
            ws.append(row_out)

    output_wb.save(config.output_path)
    workbook.close()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", dest="input_path", required=True)
    parser.add_argument("--output", dest="output_path", required=True)
    parser.add_argument("--sheet-name", dest="sheet_name", default=None)
    parser.add_argument("--window-seconds", type=int, default=5)
    parser.add_argument("--histogram-min", type=float, default=0.0)
    parser.add_argument("--histogram-max", type=float, default=200.0)
    parser.add_argument("--histogram-bin-width", type=float, default=0.5)
    parser.add_argument("--free-flow-speed", type=float, default=None)
    parser.add_argument("--free-flow-quantile", type=float, default=0.90)
    parser.add_argument("--low-speed-ratio-to-free-flow", type=float, default=0.50)
    parser.add_argument("--near-stop-threshold", type=float, default=5.0)
    parser.add_argument("--hard-brake-threshold", type=float, default=-8.0)
    parser.add_argument("--speed-drop-threshold", type=float, default=5.0)
    parser.add_argument("--short-time-headway-threshold", type=float, default=1.0)
    parser.add_argument("--min-points-per-vehicle", type=int, default=3)
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    config = Config(**vars(args))
    base_time_ms, free_flow_speed = first_pass(config)
    second_pass(config, base_time_ms=base_time_ms, free_flow_speed=free_flow_speed)


if __name__ == "__main__":
    main()
